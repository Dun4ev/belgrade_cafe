import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_model():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=11, bold=True, color="000000")
    font_bold = Font(name="Calibri", size=11, bold=True, color="000000")
    font_regular = Font(name="Calibri", size=11, color="000000")
    font_input = Font(name="Calibri", size=11, color="0000FF") # Blue for inputs
    font_formula = Font(name="Calibri", size=11, color="000000") # Black for formulas
    
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Navy
    fill_section = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Soft Blue
    fill_accent = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Light Gray
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom = Border(bottom=Side(border_style="medium", color="000000"))
    double_bottom = Border(bottom=Side(border_style="double", color="000000"), top=Side(border_style="thin", color="D3D3D3"))

    # 1. INPUTS SHEET
    ws_in = wb.create_sheet("Inputs")
    ws_in.views.sheetView[0].showGridLines = True
    
    # Headers
    ws_in.merge_cells("A1:C1")
    ws_in["A1"] = "Входные параметры (Belgrade Cafe Bar Model)"
    ws_in["A1"].font = font_title
    ws_in["A1"].fill = fill_header
    ws_in["A1"].alignment = Alignment(horizontal="center")
    ws_in.row_dimensions[1].height = 30
    
    inputs_data = [
        # Section, Parameter, Value, Unit
        ("1. Капитальные затраты (CapEx)", "", "", ""),
        ("", "Аренда помещения (депозит + 1 месяц)", 1800, "EUR"),
        ("", "Ремонт, дизайн и отделка", 8000, "EUR"),
        ("", "Профессиональное оборудование (б/у)", 3500, "EUR"),
        ("", "Мебель (столы, стулья, бар)", 3000, "EUR"),
        ("", "Посуда, инвентарь и POS-система", 2000, "EUR"),
        ("", "Регистрация (APR), лицензии, юристы", 1000, "EUR"),
        ("", "Маркетинг, вывеска и открытие", 700, "EUR"),
        ("", "Резерв оборотного капитала", 4000, "EUR"),
        ("2. Операционные расходы (OpEx / месяц)", "", "", ""),
        ("", "Аренда помещения (40 кв.м.)", 600, "EUR"),
        ("", "Коммунальные услуги и интернет", 300, "EUR"),
        ("", "Количество сотрудников (бариста/официант)", 2, "чел"),
        ("", "Зарплата 1 сотрудника (нетто)", 550, "EUR"),
        ("", "Налоги на зарплату (ставка от нетто)", 0.60, "%"),
        ("", "Бухгалтерия и обслуживание ПО", 150, "EUR"),
        ("", "Музыкальные лицензии (Sokoj/OFPS)", 50, "EUR"),
        ("", "Расходные материалы и хозтовары", 100, "EUR"),
        ("3. Доходы и продажи", "", "", ""),
        ("", "Курс EUR / RSD", 117.0, "RSD"),
        ("", "Средний чек (в RSD)", 300, "RSD"),
        ("", "Среднее количество посетителей в день", 100, "чел"),
        ("", "Дней работы в месяц", 30, "дней"),
        ("", "Себестоимость напитков (COGS %)", 0.25, "%"),
        ("", "Налог на прибыль в Сербии", 0.15, "%")
    ]
    
    row_idx = 3
    for item in inputs_data:
        sec, param, val, unit = item
        if sec:
            ws_in.cell(row=row_idx, column=1, value=sec).font = font_section
            ws_in.cell(row=row_idx, column=1).fill = fill_section
            ws_in.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
            ws_in.row_dimensions[row_idx].height = 22
            row_idx += 1
        else:
            ws_in.cell(row=row_idx, column=1, value=param).font = font_regular
            cell_val = ws_in.cell(row=row_idx, column=2, value=val)
            cell_val.font = font_input
            if unit == "%":
                cell_val.number_format = "0.0%"
            elif unit == "EUR":
                cell_val.number_format = "€#,##0"
            elif unit == "RSD":
                cell_val.number_format = "#,##0.0" if param == "Курс EUR / RSD" else "#,##0"
            else:
                cell_val.number_format = "#,##0"
            
            ws_in.cell(row=row_idx, column=3, value=unit).font = font_regular
            ws_in.cell(row=row_idx, column=1).border = thin_border
            ws_in.cell(row=row_idx, column=2).border = thin_border
            ws_in.cell(row=row_idx, column=3).border = thin_border
            row_idx += 1
            
    # Auto-fit columns for Inputs
    for col in range(1, 4):
        col_letter = get_column_letter(col)
        max_len = max(len(str(ws_in.cell(row=r, column=col).value or '')) for r in range(1, row_idx))
        ws_in.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # 2. MODEL SHEET (Calculations)
    ws_mod = wb.create_sheet("Model")
    ws_mod.views.sheetView[0].showGridLines = True
    
    ws_mod.merge_cells("A1:C1")
    ws_mod["A1"] = "Расчет окупаемости и рентабельности"
    ws_mod["A1"].font = font_title
    ws_mod["A1"].fill = fill_header
    ws_mod["A1"].alignment = Alignment(horizontal="center")
    ws_mod.row_dimensions[1].height = 30
    
    # Rows structure
    # Define cell mappings from Inputs
    # CapEx
    ws_mod["A3"] = "Общие стартовые вложения (CapEx)"
    ws_mod["A3"].font = font_bold
    ws_mod["B3"] = "=SUM(Inputs!B4:Inputs!B11)"
    ws_mod["B3"].font = font_formula
    ws_mod["B3"].number_format = "€#,##0"
    ws_mod["B3"].border = Border(bottom=Side(style='thin', color='000000'))
    ws_mod["C3"] = "EUR"
    
    # Monthly Revenue
    ws_mod["A5"] = "Ежемесячный доход"
    ws_mod["A5"].font = font_section
    ws_mod.merge_cells("A5:C5")
    ws_mod["A5"].fill = fill_section
    
    ws_mod["A6"] = "Средний чек (в EUR)"
    ws_mod["A6"].font = font_regular
    ws_mod["B6"] = "=Inputs!B23/Inputs!B22"
    ws_mod["B6"].font = font_formula
    ws_mod["B6"].number_format = "€#,##0.00"
    ws_mod["C6"] = "EUR"
    
    ws_mod["A7"] = "Выручка в месяц"
    ws_mod["A7"].font = font_bold
    ws_mod["B7"] = "=B6*Inputs!B24*Inputs!B25"
    ws_mod["B7"].font = font_formula
    ws_mod["B7"].number_format = "€#,##0"
    ws_mod["C7"] = "EUR"
    
    # COGS
    ws_mod["A8"] = "Себестоимость продуктов (COGS)"
    ws_mod["A8"].font = font_regular
    ws_mod["B8"] = "=B7*Inputs!B26"
    ws_mod["B8"].font = font_formula
    ws_mod["B8"].number_format = "€#,##0"
    ws_mod["C8"] = "EUR"
    
    ws_mod["A9"] = "Валовая прибыль (Gross Profit)"
    ws_mod["A9"].font = font_bold
    ws_mod["B9"] = "=B7-B8"
    ws_mod["B9"].font = font_formula
    ws_mod["B9"].number_format = "€#,##0"
    ws_mod["C9"] = "EUR"
    ws_mod["B9"].border = Border(top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='000000'))
    
    # OpEx
    ws_mod["A11"] = "Ежемесячные постоянные расходы (OpEx)"
    ws_mod["A11"].font = font_section
    ws_mod.merge_cells("A11:C11")
    ws_mod["A11"].fill = fill_section
    
    ws_mod["A12"] = "Аренда помещения"
    ws_mod["A12"].font = font_regular
    ws_mod["B12"] = "=Inputs!B13"
    ws_mod["B12"].font = font_formula
    ws_mod["B12"].number_format = "€#,##0"
    ws_mod["C12"] = "EUR"
    
    ws_mod["A13"] = "Коммунальные услуги"
    ws_mod["A13"].font = font_regular
    ws_mod["B13"] = "=Inputs!B14"
    ws_mod["B13"].font = font_formula
    ws_mod["B13"].number_format = "€#,##0"
    ws_mod["C13"] = "EUR"
    
    ws_mod["A14"] = "Фонд оплаты труда (ФОТ с налогами)"
    ws_mod["A14"].font = font_regular
    ws_mod["B14"] = "=Inputs!B15*Inputs!B16*(1+Inputs!B17)"
    ws_mod["B14"].font = font_formula
    ws_mod["B14"].number_format = "€#,##0"
    ws_mod["C14"] = "EUR"
    
    ws_mod["A15"] = "Бухгалтерия и ПО"
    ws_mod["A15"].font = font_regular
    ws_mod["B15"] = "=Inputs!B18"
    ws_mod["B15"].font = font_formula
    ws_mod["B15"].number_format = "€#,##0"
    ws_mod["C15"] = "EUR"
    
    ws_mod["A16"] = "Лицензии и прочие расходы"
    ws_mod["A16"].font = font_regular
    ws_mod["B16"] = "=Inputs!B19+Inputs!B20"
    ws_mod["B16"].font = font_formula
    ws_mod["B16"].number_format = "€#,##0"
    ws_mod["C16"] = "EUR"
    
    ws_mod["A17"] = "Итого постоянные расходы (Total OpEx)"
    ws_mod["A17"].font = font_bold
    ws_mod["B17"] = "=SUM(B12:B16)"
    ws_mod["B17"].font = font_formula
    ws_mod["B17"].number_format = "€#,##0"
    ws_mod["C17"] = "EUR"
    ws_mod["B17"].border = Border(top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='000000'))
    
    # Operating Profit / Profitability
    ws_mod["A19"] = "Финансовый результат"
    ws_mod["A19"].font = font_section
    ws_mod.merge_cells("A19:C19")
    ws_mod["A19"].fill = fill_section
    
    ws_mod["A20"] = "EBITDA (Прибыль до налогов)"
    ws_mod["A20"].font = font_bold
    ws_mod["B20"] = "=B9-B17"
    ws_mod["B20"].font = font_formula
    ws_mod["B20"].number_format = "€#,##0"
    ws_mod["C20"] = "EUR"
    
    ws_mod["A21"] = "Налог на прибыль (15%)"
    ws_mod["A21"].font = font_regular
    ws_mod["B21"] = "=IF(B20>0, B20*Inputs!B27, 0)"
    ws_mod["B21"].font = font_formula
    ws_mod["B21"].number_format = "€#,##0"
    ws_mod["C21"] = "EUR"
    
    ws_mod["A22"] = "Чистая прибыль в месяц (Net Profit)"
    ws_mod["A22"].font = font_bold
    ws_mod["B22"] = "=B20-B21"
    ws_mod["B22"].font = font_formula
    ws_mod["B22"].number_format = "€#,##0"
    ws_mod["C22"] = "EUR"
    ws_mod["B22"].border = double_bottom
    
    # Payback
    ws_mod["A24"] = "Срок окупаемости (Payback Period)"
    ws_mod["A24"].font = font_bold
    ws_mod["B24"] = "=B3/B22"
    ws_mod["B24"].font = font_formula
    ws_mod["B24"].number_format = "0.0"
    ws_mod["C24"] = "месяцев"
    ws_mod["B24"].border = Border(bottom=Side(style='medium', color='1F4E78'))
    
    # Breakeven Customers
    ws_mod["A26"] = "Точка безубыточности (посетителей в день)"
    ws_mod["A26"].font = font_bold
    # Breakeven formula: Fixed costs / (Average Check in EUR * (1 - COGS%)) / Days
    ws_mod["B26"] = "=B17/(B6*(1-Inputs!B26))/Inputs!B25"
    ws_mod["B26"].font = font_formula
    ws_mod["B26"].number_format = "0"
    ws_mod["C26"] = "чел/день"
    
    # Border & padding formatting for all rows
    for r in range(3, 27):
        if ws_mod.cell(row=r, column=1).value != "Ежемесячный доход" and \
           ws_mod.cell(row=r, column=1).value != "Ежемесячные постоянные расходы (OpEx)" and \
           ws_mod.cell(row=r, column=1).value != "Финансовый результат":
            ws_mod.cell(row=r, column=1).border = thin_border
            ws_mod.cell(row=r, column=2).border = thin_border
            ws_mod.cell(row=r, column=3).border = thin_border
            
    # Auto-fit columns for Model
    for col in range(1, 4):
        col_letter = get_column_letter(col)
        max_len = max(len(str(ws_mod.cell(row=r, column=col).value or '')) for r in range(1, 27))
        ws_mod.column_dimensions[col_letter].width = max(max_len + 3, 15)

    # 3. SUMMARY DASHBOARD SHEET
    ws_sum = wb.create_sheet("Summary")
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.merge_cells("A1:C1")
    ws_sum["A1"] = "Резюме бизнес-плана кафе-бара"
    ws_sum["A1"].font = font_title
    ws_sum["A1"].fill = fill_header
    ws_sum["A1"].alignment = Alignment(horizontal="center")
    ws_sum.row_dimensions[1].height = 30
    
    summary_items = [
        ("Необходимый стартовый бюджет (CapEx)", "='Model'!B3", "EUR"),
        ("Ежемесячная выручка", "='Model'!B7", "EUR"),
        ("Постоянные затраты в месяц (OpEx)", "='Model'!B17", "EUR"),
        ("Чистая прибыль в месяц", "='Model'!B22", "EUR"),
        ("Срок окупаемости проекта", "='Model'!B24", "месяцев"),
        ("Точка безубыточности (клиентов в день)", "='Model'!B26", "чел/день")
    ]
    
    for idx, (label, formula, unit) in enumerate(summary_items, start=3):
        ws_sum.cell(row=idx, column=1, value=label).font = font_bold
        ws_sum.cell(row=idx, column=1).fill = fill_accent
        cell_val = ws_sum.cell(row=idx, column=2, value=formula)
        cell_val.font = font_formula
        if "B24" in formula:
            cell_val.number_format = "0.0"
        elif "B26" in formula:
            cell_val.number_format = "0"
        else:
            cell_val.number_format = "€#,##0"
            
        ws_sum.cell(row=idx, column=3, value=unit).font = font_regular
        ws_sum.cell(row=idx, column=1).border = thin_border
        ws_sum.cell(row=idx, column=2).border = thin_border
        ws_sum.cell(row=idx, column=3).border = thin_border
        
    for col in range(1, 4):
        col_letter = get_column_letter(col)
        max_len = max(len(str(ws_sum.cell(row=r, column=col).value or '')) for r in range(1, 10))
        ws_sum.column_dimensions[col_letter].width = max(max_len + 3, 15)

    # Save
    out_dir = "/Users/j15/.gemini/antigravity/scratch"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "belgrade_cafe_model.xlsx")
    wb.save(out_path)
    print(f"Model successfully saved to {out_path}")

if __name__ == "__main__":
    create_model()
