import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_model():
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    font_title = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=11, bold=True, color="000000")
    font_bold = Font(name="Calibri", size=11, bold=True, color="000000")
    font_regular = Font(name="Calibri", size=11, color="000000")
    font_input = Font(name="Calibri", size=11, color="0000FF") # Blue for inputs
    font_formula = Font(name="Calibri", size=11, color="000000") # Black for formulas
    
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Dark Navy
    fill_section = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Muted Blue
    fill_check = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid") # Accent Blue for Checks
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom = Border(bottom=Side(border_style="medium", color="000000"))
    double_bottom = Border(bottom=Side(border_style="double", color="000000"), top=Side(border_style="thin", color="D3D3D3"))

    # 1. ASSUMPTIONS SHEET
    ws_as = wb.create_sheet("Assumptions")
    ws_as.views.sheetView[0].showGridLines = True
    
    # Headers
    ws_as.merge_cells("A1:C1")
    ws_as["A1"] = "Входные параметры (Assumptions)"
    ws_as["A1"].font = font_title
    ws_as["A1"].fill = fill_header
    ws_as["A1"].alignment = Alignment(horizontal="center")
    ws_as.row_dimensions[1].height = 30
    
    assumptions = [
        # Section, Parameter, Value, Unit
        ("1. Финансирование и Стартовые активы", "", "", ""),
        ("", "Начальный капитал (собственный)", 24000, "EUR"),
        ("", "Ремонт и отделка помещения (PP&E)", 8000, "EUR"),
        ("", "Профессиональное б/у оборудование (PP&E)", 3500, "EUR"),
        ("", "Мебель (PP&E)", 3000, "EUR"),
        ("", "Депозит за аренду (долгосрочный актив)", 1200, "EUR"),
        ("", "Первоначальный товарный запас (оборотный актив)", 2000, "EUR"),
        ("", "Разовые расходы на открытие (юристы, APR, маркетинг)", 1700, "EUR"),
        ("2. Аренда и Операционные расходы в месяц", "", "", ""),
        ("", "Аренда помещения (40 кв.м.)", 600, "EUR"),
        ("", "Коммунальные услуги", 300, "EUR"),
        ("", "Количество сотрудников", 2, "чел"),
        ("", "Зарплата 1 сотрудника (нетто)", 550, "EUR"),
        ("", "Налог на зарплату (соцотчисления сверх нетто)", 0.60, "%"),
        ("", "Бухгалтерия и ПО", 150, "EUR"),
        ("", "Прочие расходы (лицензии Sokoj/OFPS, хозтовары)", 150, "EUR"),
        ("3. Продажи и Себестоимость", "", "", ""),
        ("", "Курс EUR / RSD", 117.0, "RSD"),
        ("", "Средний чек", 300, "RSD"),
        ("", "Количество гостей в день (steady state)", 100, "чел"),
        ("", "Рабочих дней в месяце", 30, "дней"),
        ("", "Себестоимость продуктов (COGS %)", 0.25, "%"),
        ("", "Налог на прибыль в Сербии", 0.15, "%"),
        ("", "Срок полезного использования PP&E", 60, "мес."),
        ("4. Темп выхода на мощность (Ramp-up)", "", "", ""),
        ("", "Месяц 1", 0.50, "%"),
        ("", "Месяц 2", 0.70, "%"),
        ("", "Месяц 3", 0.85, "%"),
        ("", "Месяцы 4-12", 1.00, "%")
    ]
    
    row_idx = 3
    for item in assumptions:
        sec, param, val, unit = item
        if sec:
            ws_as.cell(row=row_idx, column=1, value=sec).font = font_section
            ws_as.cell(row=row_idx, column=1).fill = fill_section
            ws_as.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
            ws_as.row_dimensions[row_idx].height = 22
            row_idx += 1
        else:
            ws_as.cell(row=row_idx, column=1, value=param).font = font_regular
            cell_val = ws_as.cell(row=row_idx, column=2, value=val)
            cell_val.font = font_input
            if unit == "%":
                cell_val.number_format = "0.0%"
            elif unit in ["EUR", "RSD"] and param != "Курс EUR / RSD":
                cell_val.number_format = "€#,##0" if unit == "EUR" else "#,##0"
            else:
                cell_val.number_format = "#,##0.0" if param == "Курс EUR / RSD" else "#,##0"
                
            ws_as.cell(row=row_idx, column=3, value=unit).font = font_regular
            ws_as.cell(row=row_idx, column=1).border = thin_border
            ws_as.cell(row=row_idx, column=2).border = thin_border
            ws_as.cell(row=row_idx, column=3).border = thin_border
            row_idx += 1
            
    for col in range(1, 4):
        col_letter = get_column_letter(col)
        max_len = max(len(str(ws_as.cell(row=r, column=col).value or '')) for r in range(1, row_idx))
        ws_as.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 2. MODEL SHEET
    ws_m = wb.create_sheet("Model")
    ws_m.views.sheetView[0].showGridLines = True
    
    # Setup Columns C to O (Month 0 to Month 12)
    # Col A: Line Item, Col B: Units, Col C: Month 0, Col D: Month 1 ... Col O: Month 12, Col P: Total Y1
    headers_m = ["Статья", "Ед."] + ["М0 (Старт)"] + [f"М{i}" for i in range(1, 13)] + ["Итого Y1"]
    
    ws_m.row_dimensions[1].height = 30
    ws_m.merge_cells("A1:P1")
    ws_m["A1"] = "Интегрированная 3-Statement модель"
    ws_m["A1"].font = font_title
    ws_m["A1"].fill = fill_header
    ws_m["A1"].alignment = Alignment(horizontal="center")
    
    ws_m.row_dimensions[2].height = 24
    for idx, header in enumerate(headers_m, start=1):
        cell = ws_m.cell(row=2, column=idx, value=header)
        cell.font = font_bold
        cell.fill = fill_section
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Sections Mapping
    sections = [
        ("1. РАБОЧИЕ ПОКАЗАТЕЛИ", ["Коэффициент Ramp-up", "Количество гостей в день", "Средний чек (в EUR)"]),
        ("2. ОТЧЕТ О ПРИБЫЛЯХ И УБЫТКАХ (Income Statement)", ["Выручка", "Себестоимость (COGS)", "Валовая прибыль (Gross Profit)", "Аренда помещения", "Коммунальные услуги", "ФОТ сотрудников с налогами", "Бухгалтерия и ПО", "Прочие расходы (лицензии, расходники)", "EBITDA", "Амортизация (D&A)", "Операционная прибыль (EBIT)", "Налог на прибыль", "Чистая прибыль"]),
        ("3.БАЛАНСОВЫЙ ОТЧЕТ (Balance Sheet)", ["Денежные средства (Cash)", "Товарный запас (Inventory)", "Депозит за аренду (Deposit)", "Предоплата по аренде", "Основные средства (PP&E net)", "ИТОГО АКТИВЫ", "Собственный капитал", "Нераспределенная прибыль (RE)", "ИТОГО ПАССИВЫ + КАПИТАЛ", "Балансовая проверка (Asset - L&E)"]),
        ("4. ОТЧЕТ О ДВИЖЕНИИ ДЕНЕЖНЫХ СРЕДСТВ (Cash Flow Statement)", ["Чистая прибыль (Net Income)", "Корректировка: Амортизация (D&A)", "Изменение оборотного капитала", "Денежный поток от опер. деятельности (CFO)", "Капитальные затраты (CapEx)", "Депозиты и предоплаты", "Денежный поток от инвест. деятельности (CFI)", "Привлечение собственного капитала", "Денежный поток от фин. деятельности (CFF)", "Чистое изменение денежных средств", "Начальный остаток денежных средств", "Конечный остаток денежных средств", "Проверка остатка (CF Ending vs BS Cash)"])
    ]
    
    # We will write these rows and note the row index of each row so we can write formulas correctly.
    row_mapping = {}
    current_row = 3
    
    for section_name, items in sections:
        # Header row
        ws_m.cell(row=current_row, column=1, value=section_name).font = font_bold
        ws_m.cell(row=current_row, column=1).fill = fill_section
        ws_m.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=16)
        ws_m.row_dimensions[current_row].height = 20
        current_row += 1
        
        for item in items:
            ws_m.cell(row=current_row, column=1, value=item).font = font_regular
            row_mapping[item] = current_row
            current_row += 1
            
    # Now write the cells formulas column by column (C to O, Month 0 to Month 12)
    # C is Col 3 (Month 0), D is Col 4 (Month 1)... O is Col 15 (Month 12)
    # We will write formulas referencing the Assumptions tab.
    
    for col_idx in range(3, 16):
        m = col_idx - 3 # Month number (0 to 12)
        col_letter = get_column_letter(col_idx)
        
        # 1. Рабочие показатели
        # Ramp-up
        if m == 0:
            ws_m.cell(row=row_mapping["Коэффициент Ramp-up"], column=col_idx, value=0.0).number_format = "0.0%"
        elif m == 1:
            ws_m.cell(row=row_mapping["Коэффициент Ramp-up"], column=col_idx, value="=Assumptions!$B$28").number_format = "0.0%"
        elif m == 2:
            ws_m.cell(row=row_mapping["Коэффициент Ramp-up"], column=col_idx, value="=Assumptions!$B$29").number_format = "0.0%"
        elif m == 3:
            ws_m.cell(row=row_mapping["Коэффициент Ramp-up"], column=col_idx, value="=Assumptions!$B$30").number_format = "0.0%"
        else:
            ws_m.cell(row=row_mapping["Коэффициент Ramp-up"], column=col_idx, value="=Assumptions!$B$31").number_format = "0.0%"
            
        # Количество гостей
        ws_m.cell(row=row_mapping["Количество гостей в день"], column=col_idx, value=f"=Assumptions!$B$22*{col_letter}{row_mapping['Коэффициент Ramp-up']}").number_format = "#,##0"
        
        # Средний чек в EUR
        ws_m.cell(row=row_mapping["Средний чек (в EUR)"], column=col_idx, value="=Assumptions!$B$21/Assumptions!$B$20").number_format = "€#,##0.00"
        
        # 2. Income Statement
        if m == 0:
            ws_m.cell(row=row_mapping["Выручка"], column=col_idx, value=0).number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Себестоимость (COGS)"], column=col_idx, value=0).number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Валовая прибыль (Gross Profit)"], column=col_idx, value=0).number_format = "€#,##0"
            # Fixed expenses
            for item in ["Аренда помещения", "Коммунальные услуги", "ФОТ сотрудников с налогами", "Бухгалтерия и ПО", "Прочие расходы (лицензии, расходники)"]:
                ws_m.cell(row=row_mapping[item], column=col_idx, value=0).number_format = "€#,##0"
            ws_m.cell(row=row_mapping["EBITDA"], column=col_idx, value=0).number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Амортизация (D&A)"], column=col_idx, value=0).number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Операционная прибыль (EBIT)"], column=col_idx, value=0).number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Налог на прибыль"], column=col_idx, value=0).number_format = "€#,##0"
            # startup legal/marketing costs are expensed in Month 0
            ws_m.cell(row=row_mapping["Чистая прибыль"], column=col_idx, value="=-Assumptions!$B$10").number_format = "€#,##0"
        else:
            ws_m.cell(row=row_mapping["Выручка"], column=col_idx, value=f"={col_letter}{row_mapping['Количество гостей в день']}*{col_letter}{row_mapping['Средний чек (в EUR)']}*Assumptions!$B$23").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Себестоимость (COGS)"], column=col_idx, value=f"=-{col_letter}{row_mapping['Выручка']}*Assumptions!$B$24").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Валовая прибыль (Gross Profit)"], column=col_idx, value=f"={col_letter}{row_mapping['Выручка']}+{col_letter}{row_mapping['Себестоимость (COGS)']}").number_format = "€#,##0"
            
            # OpEx (negative signs to show expense flows in IS)
            ws_m.cell(row=row_mapping["Аренда помещения"], column=col_idx, value="=-Assumptions!$B$12").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Коммунальные услуги"], column=col_idx, value="=-Assumptions!$B$13").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["ФОТ сотрудников с налогами"], column=col_idx, value="=-Assumptions!$B$14*Assumptions!$B$15*(1+Assumptions!$B$16)").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Бухгалтерия и ПО"], column=col_idx, value="=-Assumptions!$B$17").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Прочие расходы (лицензии, расходники)"], column=col_idx, value="=-Assumptions!$B$18").number_format = "€#,##0"
            
            ws_m.cell(row=row_mapping["EBITDA"], column=col_idx, value=f"={col_letter}{row_mapping['Валовая прибыль (Gross Profit)']}+SUM({col_letter}{row_mapping['Аренда помещения']}:{col_letter}{row_mapping['Прочие расходы (лицензии, расходники)']})").number_format = "€#,##0"
            
            # D&A
            ws_m.cell(row=row_mapping["Амортизация (D&A)"], column=col_idx, value="=-(Assumptions!$B$5+Assumptions!$B$6+Assumptions!$B$7)/Assumptions!$B$26").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Операционная прибыль (EBIT)"], column=col_idx, value=f"={col_letter}{row_mapping['EBITDA']}+{col_letter}{row_mapping['Амортизация (D&A)']}").number_format = "€#,##0"
            
            # Tax
            ws_m.cell(row=row_mapping["Налог на прибыль"], column=col_idx, value=f"=IF({col_letter}{row_mapping['Операционная прибыль (EBIT)']}>0, -{col_letter}{row_mapping['Операционная прибыль (EBIT)']}*Assumptions!$B$25, 0)").number_format = "€#,##0"
            
            ws_m.cell(row=row_mapping["Чистая прибыль"], column=col_idx, value=f"={col_letter}{row_mapping['Операционная прибыль (EBIT)']}+{col_letter}{row_mapping['Налог на прибыль']}").number_format = "€#,##0"

        # 3. Balance Sheet
        if m == 0:
            # Cash: Startup Equity (24k) - Startup CapEx (8k + 3.5k + 3k) - Deposit (1.2k) - Rent Prepayment for Month 1 (600) - Inventory (2k) - Startup Exp (1.7k) = 4,000 €
            ws_m.cell(row=row_mapping["Денежные средства (Cash)"], column=col_idx, value="=Assumptions!$B$4-(Assumptions!$B$5+Assumptions!$B$6+Assumptions!$B$7)-Assumptions!$B$8-Assumptions!$B$12-Assumptions!$B$9-Assumptions!$B$10").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Товарный запас (Inventory)"], column=col_idx, value="=Assumptions!$B$9").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Депозит за аренду (Deposit)"], column=col_idx, value="=Assumptions!$B$8").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Предоплата по аренде"], column=col_idx, value="=Assumptions!$B$12").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Основные средства (PP&E net)"], column=col_idx, value="=Assumptions!$B$5+Assumptions!$B$6+Assumptions!$B$7").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["ИТОГО АКТИВЫ"], column=col_idx, value=f"=SUM({col_letter}{row_mapping['Денежные средства (Cash)']}:{col_letter}{row_mapping['Основные средства (PP&E net)']})").number_format = "€#,##0"
            
            ws_m.cell(row=row_mapping["Собственный капитал"], column=col_idx, value="=Assumptions!$B$4").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Нераспределенная прибыль (RE)"], column=col_idx, value=f"={col_letter}{row_mapping['Чистая прибыль']}").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["ИТОГО ПАССИВЫ + КАПИТАЛ"], column=col_idx, value=f"={col_letter}{row_mapping['Собственный капитал']}+{col_letter}{row_mapping['Нераспределенная прибыль (RE)']}_").number_format = "€#,##0"
            # Note: we need to replace '_' in formula or use proper sum. Let's write standard summation:
            ws_m.cell(row=row_mapping["ИТОГО ПАССИВЫ + КАПИТАЛ"], column=col_idx, value=f"={col_letter}{row_mapping['Собственный капитал']}+{col_letter}{row_mapping['Нераспределенная прибыль (RE)']}").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Балансовая проверка (Asset - L&E)"], column=col_idx, value=f"={col_letter}{row_mapping['ИТОГО АКТИВЫ']}-{col_letter}{row_mapping['ИТОГО ПАССИВЫ + КАПИТАЛ']}").number_format = "€#,##0"
        else:
            prev_col = get_column_letter(col_idx - 1)
            # Cash: Ties from Cash Flow Statement Ending Cash
            ws_m.cell(row=row_mapping["Денежные средства (Cash)"], column=col_idx, value=f"={col_letter}{row_mapping['Конечный остаток денежных средств']}").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Товарный запас (Inventory)"], column=col_idx, value=f"={prev_col}{row_mapping['Товарный запас (Inventory)']}").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Депозит за аренду (Deposit)"], column=col_idx, value=f"={prev_col}{row_mapping['Депозит за аренду (Deposit)']}").number_format = "€#,##0"
            # Rent prepayment falls to 0 in Month 1 since it's expensed
            ws_m.cell(row=row_mapping["Предоплата по аренде"], column=col_idx, value=0).number_format = "€#,##0"
            # PP&E net: Prior PP&E + Current D&A (since D&A is negative in IS)
            ws_m.cell(row=row_mapping["Основные средства (PP&E net)"], column=col_idx, value=f"={prev_col}{row_mapping['Основные средства (PP&E net)']}+{col_letter}{row_mapping['Амортизация (D&A)']}").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["ИТОГО АКТИВЫ"], column=col_idx, value=f"=SUM({col_letter}{row_mapping['Денежные средства (Cash)']}:{col_letter}{row_mapping['Основные средства (PP&E net)']})").number_format = "€#,##0"
            
            ws_m.cell(row=row_mapping["Собственный капитал"], column=col_idx, value=f"={prev_col}{row_mapping['Собственный капитал']}").number_format = "€#,##0"
            # RE: Prior RE + Current Net Income
            ws_m.cell(row=row_mapping["Нераспределенная прибыль (RE)"], column=col_idx, value=f"={prev_col}{row_mapping['Нераспределенная прибыль (RE)']}+{col_letter}{row_mapping['Чистая прибыль']}").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["ИТОГО ПАССИВЫ + КАПИТАЛ"], column=col_idx, value=f"={col_letter}{row_mapping['Собственный капитал']}+{col_letter}{row_mapping['Нераспределенная прибыль (RE)']}").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Балансовая проверка (Asset - L&E)"], column=col_idx, value=f"={col_letter}{row_mapping['ИТОГО АКТИВЫ']}-{col_letter}{row_mapping['ИТОГО ПАССИВЫ + КАПИТАЛ']}").number_format = "€#,##0"

        # 4. Cash Flow Statement
        if m == 0:
            ws_m.cell(row=row_mapping["Чистая прибыль (Net Income)"], column=col_idx, value=f"={col_letter}{row_mapping['Чистая прибыль']}").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Корректировка: Амортизация (D&A)"], column=col_idx, value=0).number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Изменение оборотного капитала"], column=col_idx, value=0).number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Денежный поток от опер. деятельности (CFO)"], column=col_idx, value=f"=SUM({col_letter}{row_mapping['Чистая прибыль (Net Income)']}:{col_letter}{row_mapping['Изменение оборотного капитала']})").number_format = "€#,##0"
            
            # CFI
            ws_m.cell(row=row_mapping["Капитальные затраты (CapEx)"], column=col_idx, value="=-(Assumptions!$B$5+Assumptions!$B$6+Assumptions!$B$7)").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Депозиты и предоплаты"], column=col_idx, value="=-(Assumptions!$B$8+Assumptions!$B$12+Assumptions!$B$9)").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Денежный поток от инвест. деятельности (CFI)"], column=col_idx, value=f"={col_letter}{row_mapping['Капитальные затраты (CapEx)']}+{col_letter}{row_mapping['Депозиты и предоплаты']}").number_format = "€#,##0"
            
            # CFF
            ws_m.cell(row=row_mapping["Привлечение собственного капитала"], column=col_idx, value="=Assumptions!$B$4").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Денежный поток от фин. деятельности (CFF)"], column=col_idx, value=f"={col_letter}{row_mapping['Привлечение собственного капитала']}").number_format = "€#,##0"
            
            # Net Cash
            ws_m.cell(row=row_mapping["Чистое изменение денежных средств"], column=col_idx, value=f"={col_letter}{row_mapping['Денежный поток от опер. деятельности (CFO)']}+{col_letter}{row_mapping['Денежный поток от инвест. деятельности (CFI)']}+{col_letter}{row_mapping['Денежный поток от фин. деятельности (CFF)']}").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Начальный остаток денежных средств"], column=col_idx, value=0).number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Конечный остаток денежных средств"], column=col_idx, value=f"={col_letter}{row_mapping['Начальный остаток денежных средств']}+{col_letter}{row_mapping['Чистое изменение денежных средств']}").number_format = "€#,##0"
            
            ws_m.cell(row=row_mapping["Проверка остатка (CF Ending vs BS Cash)"], column=col_idx, value=f"={col_letter}{row_mapping['Конечный остаток денежных средств']}-{col_letter}{row_mapping['Денежные средства (Cash)']}").number_format = "€#,##0"
        else:
            prev_col = get_column_letter(col_idx - 1)
            ws_m.cell(row=row_mapping["Чистая прибыль (Net Income)"], column=col_idx, value=f"={col_letter}{row_mapping['Чистая прибыль']}").number_format = "€#,##0"
            # D&A is added back (represented as positive, so -D&A since D&A row in IS is negative)
            ws_m.cell(row=row_mapping["Корректировка: Амортизация (D&A)"], column=col_idx, value=f"=-{col_letter}{row_mapping['Амортизация (D&A)']}").number_format = "€#,##0"
            
            # Prepaid rent is expensed in Month 1, representing a positive change in working capital cash of +600 €
            if m == 1:
                ws_m.cell(row=row_mapping["Изменение оборотного капитала"], column=col_idx, value=f"={prev_col}{row_mapping['Предоплата по аренде']}-{col_letter}{row_mapping['Предоплата по аренде']}").number_format = "€#,##0"
            else:
                ws_m.cell(row=row_mapping["Изменение оборотного капитала"], column=col_idx, value=0).number_format = "€#,##0"
                
            ws_m.cell(row=row_mapping["Денежный поток от опер. деятельности (CFO)"], column=col_idx, value=f"=SUM({col_letter}{row_mapping['Чистая прибыль (Net Income)']}:{col_letter}{row_mapping['Изменение оборотного капитала']})").number_format = "€#,##0"
            
            # CFI
            ws_m.cell(row=row_mapping["Капитальные затраты (CapEx)"], column=col_idx, value=0).number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Депозиты и предоплаты"], column=col_idx, value=0).number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Денежный поток от инвест. деятельности (CFI)"], column=col_idx, value=0).number_format = "€#,##0"
            
            # CFF
            ws_m.cell(row=row_mapping["Привлечение собственного капитала"], column=col_idx, value=0).number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Денежный поток от фин. деятельности (CFF)"], column=col_idx, value=0).number_format = "€#,##0"
            
            # Net Cash
            ws_m.cell(row=row_mapping["Чистое изменение денежных средств"], column=col_idx, value=f"={col_letter}{row_mapping['Денежный поток от опер. деятельности (CFO)']}+{col_letter}{row_mapping['Денежный поток от инвест. деятельности (CFI)']}+{col_letter}{row_mapping['Денежный поток от фин. деятельности (CFF)']}").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Начальный остаток денежных средств"], column=col_idx, value=f"={prev_col}{row_mapping['Конечный остаток денежных средств']}").number_format = "€#,##0"
            ws_m.cell(row=row_mapping["Конечный остаток денежных средств"], column=col_idx, value=f"={col_letter}{row_mapping['Начальный остаток денежных средств']}+{col_letter}{row_mapping['Чистое изменение денежных средств']}").number_format = "€#,##0"
            
            ws_m.cell(row=row_mapping["Проверка остатка (CF Ending vs BS Cash)"], column=col_idx, value=f"={col_letter}{row_mapping['Конечный остаток денежных средств']}-{col_letter}{row_mapping['Денежные средства (Cash)']}").number_format = "€#,##0"

    # Write Total Column (P, Column 16)
    col_letter_t = "P"
    ws_m.cell(row=row_mapping["Коэффициент Ramp-up"], column=16, value="-").alignment = Alignment(horizontal="center")
    ws_m.cell(row=row_mapping["Количество гостей в день"], column=16, value=f"=AVERAGE(D{row_mapping['Количество гостей в день']}:O{row_mapping['Количество гостей в день']})").number_format = "#,##0"
    ws_m.cell(row=row_mapping["Средний чек (в EUR)"], column=16, value=f"=AVERAGE(D{row_mapping['Средний чек (в EUR)']}:O{row_mapping['Средний чек (в EUR)']})").number_format = "€#,##0.00"
    
    # IS Totals
    ws_m.cell(row=row_mapping["Выручка"], column=16, value=f"=SUM(D{row_mapping['Выручка']}:O{row_mapping['Выручка']})").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Себестоимость (COGS)"], column=16, value=f"=SUM(D{row_mapping['Себестоимость (COGS)']}:O{row_mapping['Себестоимость (COGS)']})").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Валовая прибыль (Gross Profit)"], column=16, value=f"=SUM(D{row_mapping['Валовая прибыль (Gross Profit)']}:O{row_mapping['Валовая прибыль (Gross Profit)']})").number_format = "€#,##0"
    
    for item in ["Аренда помещения", "Коммунальные услуги", "ФОТ сотрудников с налогами", "Бухгалтерия и ПО", "Прочие расходы (лицензии, расходники)"]:
        r = row_mapping[item]
        ws_m.cell(row=r, column=16, value=f"=SUM(D{r}:O{r})").number_format = "€#,##0"
        
    ws_m.cell(row=row_mapping["EBITDA"], column=16, value=f"=SUM(D{row_mapping['EBITDA']}:O{row_mapping['EBITDA']})").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Амортизация (D&A)"], column=16, value=f"=SUM(D{row_mapping['Амортизация (D&A)']}:O{row_mapping['Амортизация (D&A)']})").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Операционная прибыль (EBIT)"], column=16, value=f"=SUM(D{row_mapping['Операционная прибыль (EBIT)']}:O{row_mapping['Операционная прибыль (EBIT)']})").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Налог на прибыль"], column=16, value=f"=SUM(D{row_mapping['Налог на прибыль']}:O{row_mapping['Налог на прибыль']})").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Чистая прибыль"], column=16, value=f"=SUM(C{row_mapping['Чистая прибыль']}:O{row_mapping['Чистая прибыль']})").number_format = "€#,##0" # Include Month 0 startup expense in Y1 total net profit
    
    # BS is ending point (from Col O)
    for item in ["Денежные средства (Cash)", "Товарный запас (Inventory)", "Депозит за аренду (Deposit)", "Предоплата по аренде", "Основные средства (PP&E net)", "ИТОГО АКТИВЫ", "Собственный капитал", "Нераспределенная прибыль (RE)", "ИТОГО ПАССИВЫ + КАПИТАЛ", "Балансовая проверка (Asset - L&E)"]:
        r = row_mapping[item]
        ws_m.cell(row=r, column=16, value=f"=O{r}").number_format = "€#,##0"
 
    # CF Statement totals
    ws_m.cell(row=row_mapping["Чистая прибыль (Net Income)"], column=16, value=f"=SUM(C{row_mapping['Чистая прибыль (Net Income)']}:O{row_mapping['Чистая прибыль (Net Income)']})").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Корректировка: Амортизация (D&A)"], column=16, value=f"=SUM(D{row_mapping['Корректировка: Амортизация (D&A)']}:O{row_mapping['Корректировка: Амортизация (D&A)']})").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Изменение оборотного капитала"], column=16, value=f"=SUM(D{row_mapping['Изменение оборотного капитала']}:O{row_mapping['Изменение оборотного капитала']})").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Денежный поток от опер. деятельности (CFO)"], column=16, value=f"=SUM(C{row_mapping['Денежный поток от опер. деятельности (CFO)']}:O{row_mapping['Денежный поток от опер. деятельности (CFO)']})").number_format = "€#,##0"
    
    ws_m.cell(row=row_mapping["Капитальные затраты (CapEx)"], column=16, value=f"=SUM(C{row_mapping['Капитальные затраты (CapEx)']}:O{row_mapping['Капитальные затраты (CapEx)']})").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Депозиты и предоплаты"], column=16, value=f"=SUM(C{row_mapping['Депозиты и предоплаты']}:O{row_mapping['Депозиты и предоплаты']})").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Денежный поток от инвест. деятельности (CFI)"], column=16, value=f"=SUM(C{row_mapping['Денежный поток от инвест. деятельности (CFI)']}:O{row_mapping['Денежный поток от инвест. деятельности (CFI)']})").number_format = "€#,##0"
    
    ws_m.cell(row=row_mapping["Привлечение собственного капитала"], column=16, value=f"=SUM(C{row_mapping['Привлечение собственного капитала']}:O{row_mapping['Привлечение собственного капитала']})").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Денежный поток от фин. деятельности (CFF)"], column=16, value=f"=SUM(C{row_mapping['Денежный поток от фин. деятельности (CFF)']}:O{row_mapping['Денежный поток от фин. деятельности (CFF)']})").number_format = "€#,##0"
    
    ws_m.cell(row=row_mapping["Чистое изменение денежных средств"], column=16, value=f"=SUM(C{row_mapping['Чистое изменение денежных средств']}:O{row_mapping['Чистое изменение денежных средств']})").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Начальный остаток денежных средств"], column=16, value=f"=C{row_mapping['Начальный остаток денежных средств']}").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Конечный остаток денежных средств"], column=16, value=f"=O{row_mapping['Конечный остаток денежных средств']}").number_format = "€#,##0"
    ws_m.cell(row=row_mapping["Проверка остатка (CF Ending vs BS Cash)"], column=16, value=f"=O{row_mapping['Проверка остатка (CF Ending vs BS Cash)']}").number_format = "€#,##0"

    # Style cells and add borders
    for r in range(3, current_row):
        is_header = ws_m.cell(row=r, column=1).value in [s[0] for s in sections]
        for c in range(1, 17):
            cell = ws_m.cell(row=r, column=c)
            if not is_header:
                cell.border = thin_border
            
            # Double underlines for key checks & totals
            if r in [row_mapping["ИТОГО АКТИВЫ"], row_mapping["ИТОГО ПАССИВЫ + КАПИТАЛ"], row_mapping["Конечный остаток денежных средств"]]:
                cell.font = font_bold
                cell.border = Border(top=Side(style='thin', color='D3D3D3'), bottom=Side(style='double', color='000000'))
                
            if r in [row_mapping["Балансовая проверка (Asset - L&E)"], row_mapping["Проверка остатка (CF Ending vs BS Cash)"]]:
                cell.font = font_bold
                cell.fill = fill_check
                
    # Auto-fit columns for Model
    ws_m.column_dimensions["A"].width = 42
    ws_m.column_dimensions["B"].width = 10
    for col in range(3, 17):
        col_letter = get_column_letter(col)
        ws_m.column_dimensions[col_letter].width = 14

    # Save
    out_dir = "/Users/j15/.gemini/antigravity/scratch/belgrade_cafe"
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "belgrade_cafe_model.xlsx")
    wb.save(out_path)
    print(f"3-statement model successfully saved to {out_path}")

if __name__ == "__main__":
    create_model()
